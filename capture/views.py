from django.contrib.auth.decorators import login_required, user_passes_test
from django.views.decorators.csrf import csrf_protect
from django.core.files.base import ContentFile
from django.shortcuts import get_object_or_404, get_list_or_404, render, redirect # ✅ Ensure both are imported
from django.forms import ValidationError
from django.http import JsonResponse, HttpResponse
from django.db.models import Count, Max
from .models import CapturedImage, DeletedCapturedImage, UploadedFile, UploadedImage, IssueReport
import zipfile
import base64
import uuid
import os
import traceback  # ✅ Import for debugging
import shutil
from django.utils.timezone import now
from django.conf import settings
from members.models import CustomTag
from itertools import chain
from operator import attrgetter
import openpyxl
from datetime import datetime
from django.utils import timezone
from django.contrib import messages
from .forms import IssueReportForm, IssueAdminUpdateForm
from django.core.mail import send_mail
from django.core.exceptions import ValidationError
import traceback

# Only allow superusers to approve other users
def is_super_user(user):
    return user.is_superuser

@login_required
def capture_uhid_camera(request):
    # Define keypad layout
    keypad_rows = [
        [1, 2, 3],
        [4, 5, 6],
        [7, 8, 9],
        ["⌫", 0, "✔"]
    ]

    # Calculate last two digits of the current year
    yearPrefix = str(datetime.now().year)[-2:]
    
    # Debug log
    print(f"[DEBUG] Year Prefix calculated: {yearPrefix}")

    return render(request, "capture/capture_uhid_camera.html", {
        "user": request.user,
        "keypad_rows": keypad_rows,
        "yearPrefix": yearPrefix,  # Send to template
    })



@login_required
def capture_images(request):
    uhid = request.GET.get('uhid')

    # Fetch all available custom tags
    custom_tags = CustomTag.objects.filter(is_deleted=False).order_by('name')

    if request.method == "POST":
        image_data = request.POST.get("image_data")
        uhid = request.POST.get("uhid")
        custom_tag = request.POST.get("custom_tag")
        image_size = request.POST.get("image_size")
        latitude = request.POST.get("latitude")
        longitude = request.POST.get("longitude")

        print(f"Received UHID: {uhid}")
        print(f"Received Image Data: {image_data[:50]}...")
        print(f"Received Custom Tag: {custom_tag}")
        print(f"Received Image Size: {image_size}")
        print(f"Received Latitude: {latitude}")
        print(f"Received Longitude: {longitude}")

        try:
            # Ensure UHID is a valid integer and within range
            uhid_int = int(uhid)
            if uhid_int > 2147483647 or uhid_int < -2147483648:  # PostgreSQL Integer Range
                raise ValidationError("UHID is too large or invalid!")

            # Process and save image
            format, imgstr = image_data.split(";base64,")
            ext = format.split("/")[-1]
            file_name = f"{uuid.uuid4().hex}.{ext}"
            image_file = ContentFile(base64.b64decode(imgstr), name=file_name)

            # Handle the custom tag
            tag_obj = None
            if custom_tag:
                tag_obj = CustomTag.objects.get(id=custom_tag)

            CapturedImage.objects.create(
                user=request.user,
                uhid=uhid_int,
                image_path=image_file,
                custom_tag=tag_obj,
                image_size=image_size,
                latitude=latitude,
                longitude=longitude
                )

            print("Image successfully saved to database.")
            return JsonResponse({"success": True, "message": "Image uploaded successfully!"})

        except ValueError:
            return JsonResponse({"success": False, "error": "UHID must be a valid number!"})
        except ValidationError as e:
            return JsonResponse({"success": False, "error": str(e)})
        except Exception as e:
            print("❌ ERROR OCCURRED:", traceback.format_exc())
            return JsonResponse({"success": False, "error": "Something went wrong. Please try again."})

    return render(request, "capture/capture_images.html", {"uhid": uhid, "custom_tags": custom_tags})

MAX_FILE_SIZE = 45 * 1024 * 1024  # 45 MB limit

@login_required
def upload_file(request):
    uhid = request.GET.get("uhid")
    custom_tags = CustomTag.objects.filter(is_deleted=False).order_by('name')

    if request.method == "POST":
        custom_tag_id = request.POST.get("custom-tag-select")  
        uploaded_file = request.FILES.get("file")

        if not custom_tag_id:
            return JsonResponse({"success": False, "error": "Custom tag is required!"})

        if not uploaded_file:
            return JsonResponse({"success": False, "error": "No file uploaded!"})

        file_size = uploaded_file.size
        file_type = uploaded_file.content_type

        print(f"Received UHID (from GET): {uhid}")
        print(f"Received File: {uploaded_file}")
        print(f"Received Custom Tag: {custom_tag_id}")
        print(f"File Size: {file_size} bytes")
        print(f"File Type: {file_type}")

        try:
            uhid_int = int(uhid)
            if uhid_int > 2147483647 or uhid_int < -2147483648:
                raise ValidationError("UHID is too large or invalid!")

            if file_size > MAX_FILE_SIZE:
                raise ValidationError(f"File size exceeds the {MAX_FILE_SIZE // (1024 * 1024)}MB limit!")

            allowed_file_types = [
                'application/pdf',
                'image/jpeg',
                'image/png',
                'image/gif'
            ]
            if file_type not in allowed_file_types:
                raise ValidationError("Unsupported file type! Allowed types: PDF, JPEG, PNG, GIF.")

            # ✅ Lookup tag by ID if present
            tag_obj = None
            if custom_tag_id:
                try:
                    tag_obj = CustomTag.objects.get(id=custom_tag_id, is_deleted=False)
                except CustomTag.DoesNotExist:
                    raise ValidationError("Invalid custom tag selected.")

            UploadedFile.objects.create(
                user=request.user,
                uhid=uhid_int,
                file_path=uploaded_file,
                custom_tag=tag_obj,
                file_size=file_size,
                file_type=file_type
            )

            print("✅ File successfully uploaded.")
            return JsonResponse({"success": True, "message": "File uploaded successfully!"})

        except ValueError:
            return JsonResponse({"success": False, "error": "UHID must be a valid number!"})
        except ValidationError as e:
            return JsonResponse({"success": False, "error": str(e)})
        except Exception as e:
            print("❌ ERROR OCCURRED:", traceback.format_exc())
            return JsonResponse({"success": False, "error": "Something went wrong. Please try again."})

    return render(request, "capture/upload_file.html", {
        "uhid": uhid,
        "custom_tags": custom_tags
    })

@login_required
def upload_file_image2pdf(request):
    uhid = request.GET.get("uhid")
    custom_tags = CustomTag.objects.filter(is_deleted=False).order_by('name')

    if request.method == "POST":
        custom_tag_id = request.POST.get("custom-tag-select")
        uploaded_file = request.FILES.get("file")

        if not custom_tag_id:
            return JsonResponse({"success": False, "error": "Custom tag is required!"})

        if not uploaded_file:
            return JsonResponse({"success": False, "error": "No file uploaded!"})

        file_size = uploaded_file.size
        file_type = uploaded_file.content_type

        print(f"Received UHID (from GET): {uhid}")
        print(f"Received File: {uploaded_file}")
        print(f"Received Custom Tag: {custom_tag_id}")
        print(f"File Size: {file_size} bytes")
        print(f"File Type: {file_type}")

        try:
            uhid_int = int(uhid)
            if uhid_int > 2147483647 or uhid_int < -2147483648:
                raise ValidationError("UHID is too large or invalid!")

            if file_size > MAX_FILE_SIZE:
                raise ValidationError(f"File size exceeds the {MAX_FILE_SIZE // (1024 * 1024)}MB limit!")

            allowed_file_types = [
                'application/pdf',
                'image/jpeg',
                'image/jpg',
                'image/png'
            ]
            if file_type not in allowed_file_types:
                raise ValidationError("Unsupported file type! Allowed types: PDF, JPEG, PNG.")

            # ✅ Lookup tag by ID if present
            tag_obj = None
            if custom_tag_id:
                try:
                    tag_obj = CustomTag.objects.get(id=custom_tag_id, is_deleted=False)
                except CustomTag.DoesNotExist:
                    raise ValidationError("Invalid custom tag selected.")

            # Save the uploaded file to the database
            UploadedFile.objects.create(
                user=request.user,
                uhid=uhid_int,
                file_path=uploaded_file,
                custom_tag=tag_obj,
                file_size=file_size,
                file_type=file_type
            )

            print("✅ File successfully uploaded.")
            return JsonResponse({"success": True, "message": "File uploaded successfully!"})

        except ValueError:
            return JsonResponse({"success": False, "error": "UHID must be a valid number!"})
        except ValidationError as e:
            return JsonResponse({"success": False, "error": str(e)})
        except Exception as e:
            print("❌ ERROR OCCURRED:", traceback.format_exc())
            return JsonResponse({"success": False, "error": "Something went wrong. Please try again."})

    return render(request, "capture/upload_file_image2pdf.html", {
        "uhid": uhid,
        "custom_tags": custom_tags
    })


@login_required
def upload_file_pdf2pdf(request):
    uhid = request.GET.get("uhid")
    custom_tags = CustomTag.objects.filter(is_deleted=False).order_by('name')

    if request.method == "POST":
        custom_tag_id = request.POST.get("custom-tag-select")
        uploaded_files = request.FILES.getlist("files")  # Get a list of PDFs :contentReference[oaicite:26]{index=26}

        if not custom_tag_id:
            return JsonResponse({"success": False, "error": "Custom tag is required!"})

        if not uploaded_files:
            return JsonResponse({"success": False, "error": "No files uploaded!"})

        try:
            # Validate UHID once
            uhid_int = int(uhid)
            if uhid_int > 2147483647 or uhid_int < -2147483648:
                raise ValidationError("UHID is too large or invalid!")
        except ValueError:
            return JsonResponse({"success": False, "error": "UHID must be a valid number!"})
        except ValidationError as e:
            return JsonResponse({"success": False, "error": str(e)})

        # Validate custom tag
        try:
            tag_obj = CustomTag.objects.get(id=custom_tag_id, is_deleted=False)
        except CustomTag.DoesNotExist:
            return JsonResponse({"success": False, "error": "Invalid custom tag selected."})

        # Process each PDF file individually
        results = []
        for pdf in uploaded_files:
            file_name = pdf.name
            file_size = pdf.size
            file_type = pdf.content_type

            print(f"Received UHID (from GET): {uhid}")
            print(f"Received File: {file_name}")
            print(f"Received Custom Tag: {custom_tag_id}")
            print(f"File Size: {file_size} bytes")
            print(f"File Type: {file_type}")

            try:
                # Validate size
                if file_size > MAX_FILE_SIZE:
                    raise ValidationError(f"\"{file_name}\" exceeds the {MAX_FILE_SIZE // (1024 * 1024)}MB limit!")

                # Validate MIME type is PDF
                if file_type != "application/pdf":
                    raise ValidationError(f"\"{file_name}\" unsupported type! Only PDF files are allowed.")

                # Save this PDF to the UploadedFile model
                UploadedFile.objects.create(
                    user=request.user,
                    uhid=uhid_int,
                    file_path=pdf,  # Django will handle storing the PDF
                    custom_tag=tag_obj,
                    file_size=file_size,
                    file_type=file_type
                )

                print(f"✅ \"{file_name}\" uploaded successfully.")
                results.append({"file": file_name, "success": True})

            except ValidationError as e:
                print(f"❌ Validation error for {file_name}: {e}")
                results.append({"file": file_name, "success": False, "error": str(e)})
                # Continue to attempt saving other files

            except Exception:
                # Log the traceback for debugging
                print("❌ ERROR OCCURRED:", traceback.format_exc())
                results.append({"file": file_name, "success": False, "error": "Unexpected server error"})

        # Determine overall success only if every PDF succeeded
        all_success = all(r.get("success") for r in results)
        return JsonResponse({"success": all_success, "results": results})

    # GET: render the upload page
    return render(request, "capture/upload_file_pdf2pdf.html", {
        "uhid": uhid,
        "custom_tags": custom_tags
    })


@login_required
def upload_file_image2image(request):
    uhid = request.GET.get("uhid")
    custom_tags = CustomTag.objects.filter(is_deleted=False).order_by('name')

    if request.method == "POST":
        custom_tag_id = request.POST.get("custom-tag-select")
        uploaded_file = request.FILES.get("file")

        if not custom_tag_id:
            return JsonResponse({"success": False, "error": "Custom tag is required!"})

        if not uploaded_file:
            return JsonResponse({"success": False, "error": "No file uploaded!"})

        file_size = uploaded_file.size
        file_type = uploaded_file.content_type

        print(f"Received UHID (from GET): {uhid}")
        print(f"Received File: {uploaded_file}")
        print(f"Received Custom Tag: {custom_tag_id}")
        print(f"File Size: {file_size} bytes")
        print(f"File Type: {file_type}")

        try:
            # Validate UHID
            uhid_int = int(uhid)
            if uhid_int > 2147483647 or uhid_int < -2147483648:
                raise ValidationError("UHID is out of valid range!")

            # Validate file size
            if file_size > MAX_FILE_SIZE:
                raise ValidationError(f"File size exceeds the {MAX_FILE_SIZE // (1024 * 1024)}MB limit!")

            # Validate file type
            allowed_file_types = ['image/jpeg', 'image/jpg', 'image/png']
            if file_type not in allowed_file_types:
                raise ValidationError("Unsupported file type! Only JPEG and PNG images are allowed.")

            # Validate custom tag
            try:
                tag_obj = CustomTag.objects.get(id=custom_tag_id, is_deleted=False)
            except CustomTag.DoesNotExist:
                raise ValidationError("Invalid custom tag selected.")

            # Save the uploaded file
            UploadedImage.objects.create(
                user=request.user,
                uhid=uhid_int,
                image_path=uploaded_file,
                custom_tag=tag_obj,
                image_size=file_size,
            )

            print("✅ Image successfully uploaded.")
            return JsonResponse({"success": True, "message": "Image uploaded successfully!"})

        except ValueError:
            return JsonResponse({"success": False, "error": "UHID must be a valid number!"})
        except ValidationError as e:
            return JsonResponse({"success": False, "error": str(e)})
        except Exception:
            print("❌ ERROR OCCURRED:", traceback.format_exc())
            return JsonResponse({"success": False, "error": "Something went wrong. Please try again."})

    return render(request, "capture/upload_file_image2image.html", {
        "uhid": uhid,
        "custom_tags": custom_tags
    })

@login_required
def upload_file_multi_image2image(request):
    uhid = request.GET.get("uhid")
    custom_tags = CustomTag.objects.filter(is_deleted=False).order_by('name')

    if request.method == "POST":
        custom_tag_id = request.POST.get("custom-tag-select")
        uploaded_files = request.FILES.getlist("files")  # <-- changed here

        # If no tag selected, return immediately
        if not custom_tag_id:
            return JsonResponse({"success": False, "error": "Custom tag is required!"})

        # If no files uploaded at all, error
        if not uploaded_files:
            return JsonResponse({"success": False, "error": "No files uploaded!"})

        # Attempt to fetch/validate the tag object once
        try:
            tag_obj = CustomTag.objects.get(id=custom_tag_id, is_deleted=False)
        except CustomTag.DoesNotExist:
            return JsonResponse({"success": False, "error": "Invalid custom tag selected."})

        results = []  # collect per-file status

        for uploaded_file in uploaded_files:
            file_name = uploaded_file.name
            file_size = uploaded_file.size
            file_type = uploaded_file.content_type

            print(f"Received UHID (from GET): {uhid}")
            print(f"Received File: {file_name}")
            print(f"Received Custom Tag: {custom_tag_id}")
            print(f"File Size: {file_size} bytes")
            print(f"File Type: {file_type}")

            try:
                # Validate UHID once per request (same for all files)
                uhid_int = int(uhid)
                if uhid_int > 2147483647 or uhid_int < 0:
                    raise ValidationError("UHID is out of valid range!")

                # Validate each file’s size
                if file_size > MAX_FILE_SIZE:
                    raise ValidationError(f"\"{file_name}\" exceeds the {MAX_FILE_SIZE // (1024 * 1024)}MB limit!")

                # Validate each file’s MIME type
                allowed_file_types = ['image/jpeg', 'image/jpg', 'image/png']
                if file_type not in allowed_file_types:
                    raise ValidationError(f"\"{file_name}\" unsupported type! Only JPEG/PNG allowed.")

                # Save this file to the UploadedImage model
                UploadedImage.objects.create(
                    user=request.user,
                    uhid=uhid_int,
                    image_path=uploaded_file,  # Django will assign a unique name under your upload_to
                    custom_tag=tag_obj,
                    image_size=file_size,
                )

                print(f"✅ \"{file_name}\" uploaded successfully.")
                results.append({"file": file_name, "success": True})

            except ValueError:
                results.append({"file": file_name, "success": False, "error": "UHID must be a valid number!"})
                # If UHID itself is invalid, break out (all files share same UHID)
                break

            except ValidationError as e:
                results.append({"file": file_name, "success": False, "error": str(e)})
                # Continue to next file—others might still succeed
                continue

            except Exception:
                # Log the traceback, add a generic error for this file
                print("❌ ERROR OCCURRED:", traceback.format_exc())
                results.append({"file": file_name, "success": False, "error": "Unexpected error. Please try again."})
                continue

        # Decide how to report back:
        # - If every entry in results has success=True, overall success.
        # - If any entry has success=False, include those errors.
        all_success = all(r.get("success") for r in results)

        return JsonResponse({
            "success": all_success,
            "results": results
        })

    # GET: render the template as before
    return render(request, "capture/upload_file_multi_image2image.html", {
        "uhid": uhid,
        "custom_tags": custom_tags
    })





@login_required
def view_images_home_uhid(request, uhid):
    """Retrieve and display images from both CapturedImage and UploadedImage models for a specific UHID."""

    # Fetch and optimize related fields
    captured_images = CapturedImage.objects.filter(uhid=uhid, is_deleted=False).select_related(
        'user__designation', 'user__department'
    ).order_by('-timestamp')

    uploaded_images = UploadedImage.objects.filter(uhid=uhid, is_deleted=False).select_related(
        'user__designation', 'user__department'
    ).order_by('-timestamp')

    # Add metadata and tag image type
    for image in captured_images:
        image.source_type = 'captured'

    for image in uploaded_images:
        image.source_type = 'uploaded'

    # Merge and sort
    all_images = sorted(
        chain(captured_images, uploaded_images),
        key=attrgetter('timestamp'),
        reverse=True
    )

    for image in all_images:
        # Calculate size in KB and MB
        try:
            if image.image_size:
                size = int(image.image_size)
                image.image_size_kb = round(size / 1024, 2)
                image.image_size_mb = round(image.image_size_kb / 1024, 2)
            else:
                image.image_size_kb = None
                image.image_size_mb = None
        except Exception as e:
            image.image_size_kb = None
            image.image_size_mb = None
            print(f"[ERROR] Invalid image size for image ID {image.id}: {e}")

        # Extract filename
        image.filename = os.path.basename(image.image_path.name)

        # Add location info for captured images
        if isinstance(image, CapturedImage):
            image.location = f"{image.latitude}, {image.longitude}" if image.latitude and image.longitude else "Unknown Location"
        else:
            image.location = "Uploaded (no location)"

        # Full name
        image.user_full_name = getattr(image.user, 'full_name', None) or "Unknown User"

        # Designation and Department (safe fallback)
        image.user_designation = getattr(image.user.designation, 'title', None)
        image.user_designation_other = getattr(image.user, 'designation_other', None)
        image.user_department = getattr(image.user.department, 'name', None)
        image.user_department_other = getattr(image.user, 'department_other', None)

        # 🔍 Debug Logging
        print("=" * 50)
        print(f"Image ID: {image.id} | Source: {image.source_type.upper()}")
        print(f"User: {image.user_full_name}")
        print(f"Designation: {image.user_designation or ''} ({image.user_designation_other or ''})")
        print(f"Department: {image.user_department or ''} ({image.user_department_other or ''})")
        print(f"Location: {image.location}")
        print(f"Size: {image.image_size_kb} KB | {image.image_size_mb} MB")
        print(f"Filename: {image.filename}")
        print("=" * 50)

    return render(request, 'capture/view_images.html', {
        'uhid': uhid,
        'images': all_images
    })




def view_files_home_uhid(request, uhid):
    # Fetch uploaded files for the specified UHID, ordered by timestamp
    files = UploadedFile.objects.filter(uhid=uhid, is_deleted=False).order_by('-timestamp')

    for file in files:
        # Calculate sizes
        file.file_size_kb = round(file.file_path.size / 1024, 2)  # Convert to KB
        file.file_size_mb = round(file.file_path.size / (1024 * 1024), 2)  # Convert to MB

        # Convert timestamp to local timezone and format nicely: "12-May-2025 10:15 AM"
        local_timestamp = timezone.localtime(file.timestamp)
        file.formatted_datetime = local_timestamp.strftime('%d-%b-%Y %I:%M %p')

        # Debugging outputs
        print(f"📝 Filename: {file.filename}")
        print(f"🌐 File URL: {file.file_path.url}")
        print(f"📂 File Path: {file.file_path.path}")
        
        # Check if file actually exists on disk
        if os.path.exists(file.file_path.path):
            print(f"✅ File exists: {file.file_path.path}")
        else:
            print(f"❌ Missing file: {file.file_path.path}")

    return render(request, 'capture/view_files.html', {
        'files': files,
        'uhid': uhid
    })


@login_required
def download_all_images(request, uhid):
    """
    Creates a ZIP file with all images (captured + uploaded) for the given UHID and returns it.
    """

    # Fetch both captured and uploaded images
    captured_images = CapturedImage.objects.filter(uhid=uhid, is_deleted=False)
    uploaded_images = UploadedImage.objects.filter(uhid=uhid, is_deleted=False)

    # Merge and sort by timestamp descending
    all_images = sorted(
        chain(captured_images, uploaded_images),
        key=attrgetter('timestamp'),
        reverse=True
    )

    # Prepare ZIP response
    zip_filename = f"UHID_{uhid}_images.zip"
    response = HttpResponse(content_type="application/zip")
    response["Content-Disposition"] = f'attachment; filename="{zip_filename}"'

    with zipfile.ZipFile(response, "w", zipfile.ZIP_DEFLATED) as zip_file:
        for image in all_images:
            if image.image_path:  # Ensure path exists
                try:
                    file_path = image.image_path.path  # Absolute path
                    filename = os.path.basename(file_path) # Original filename

                    # Write the file to the ZIP archive with its original name
                    zip_file.write(file_path, arcname=filename)
                except Exception as e:
                    print(f"[ERROR] Could not add image ID {image.id} to ZIP: {e}")

    return response


@login_required
def download_all_files(request, uhid):
    """Creates a ZIP file with all uploaded files for the given UHID and returns it."""
    
    files = get_list_or_404(UploadedFile, uhid=uhid, is_deleted=False)  # Get all files for UHID

    zip_filename = f"UHID_{uhid}_files.zip"
    response = HttpResponse(content_type="application/zip")
    response["Content-Disposition"] = f'attachment; filename="{zip_filename}"'

    with zipfile.ZipFile(response, "w", zipfile.ZIP_DEFLATED) as zip_file:
        for file in files:
            if file.file_path:  # Ensure file exists
                file_path = file.file_path.path  # Absolute file path
                filename = os.path.basename(file_path)  # Retain original filename
                
                zip_file.write(file_path, filename)

    return response


def uhid_options(request):
    uhid = request.GET.get('uhid')
    if not uhid:
        return redirect('uhid_capture_camera')  # fallback

    # Count captured images for UHID
    captured_count = CapturedImage.objects.filter(uhid=uhid, is_deleted=False).count()

    # Count uploaded images for UHID
    uploaded_image_count = UploadedImage.objects.filter(uhid=uhid, is_deleted=False).count()

    # Total image count = captured + uploaded
    image_count = captured_count + uploaded_image_count

    # Count uploaded non-image files (if applicable)
    file_count = UploadedFile.objects.filter(uhid=uhid, is_deleted=False).count()

    return render(request, 'capture/uhid_options.html', {
        'uhid': uhid,
        'image_count': image_count,
        'file_count': file_count,
    })


@login_required
def delete_image(request, model, id):

    if model == "captured":
        obj_model = CapturedImage
    elif model == "uploaded":
        obj_model = UploadedImage
    else:
        messages.error(request, "Invalid model.")
        return redirect(request.META.get("HTTP_REFERER", "/"))

    image = get_object_or_404(obj_model, id=id)

    if image.is_deleted:
        messages.warning(request, "Image already deleted.")
        return redirect(request.META.get("HTTP_REFERER", "/"))

    # Move physical file
    if image.image_path and os.path.exists(image.image_path.path):

        old_path = image.image_path.path

        deleted_folder = os.path.join(
            settings.MEDIA_ROOT,
            f"UHID_{image.uhid}",
            f"UHID_{image.uhid}_deleted_files"
        )

        os.makedirs(deleted_folder, exist_ok=True)

        new_path = os.path.join(
            deleted_folder,
            os.path.basename(old_path)
        )

        shutil.move(old_path, new_path)

        image.image_path.name = os.path.relpath(
            new_path,
            settings.MEDIA_ROOT
        )

        relative_path = os.path.relpath(new_path, settings.MEDIA_ROOT)
        relative_path = relative_path.replace("\\", "/")   # <<< FIX HERE

        image.image_path.name = relative_path

    # Soft delete
    image.is_deleted = True
    image.deleted_on = timezone.now()
    image.deleted_by = request.user
    image.save()

    messages.success(request, "Image successfully deleted")
    return redirect(request.META.get("HTTP_REFERER", "/"))


@login_required
def delete_uploaded_pdf(request, id):
    file = get_object_or_404(UploadedFile, id=id)

    # ❌ Prevent double delete
    if file.is_deleted:
        messages.warning(request, "PDF file already deleted.")
        return redirect(request.META.get("HTTP_REFERER", "/"))

    # -----------------------------
    # FILE MOVE LOGIC
    # -----------------------------
    if file.file_path and os.path.exists(file.file_path.path):
        old_path = file.file_path.path

        # New deleted folder
        deleted_folder = os.path.join(
            settings.MEDIA_ROOT,
            f"UHID_{file.uhid}",
            f"UHID_{file.uhid}_deleted_files"
        )

        os.makedirs(deleted_folder, exist_ok=True)

        new_path = os.path.join(
            deleted_folder,
            os.path.basename(old_path)
        )

        shutil.move(old_path, new_path)

        # Convert OS path → relative path → URL format
        relative_path = os.path.relpath(new_path, settings.MEDIA_ROOT)
        relative_path = relative_path.replace("\\", "/")   # <<< FIX HERE

        file.file_path.name = relative_path



    # -----------------------------
    # SOFT DELETE FIELDS
    # -----------------------------
    file.is_deleted = True
    file.deleted_on = timezone.now()
    file.deleted_by = request.user
    file.save()

    messages.success(request, "PDF file deleted successfully.")
    return redirect(request.META.get("HTTP_REFERER", "/"))



@login_required
@user_passes_test(is_super_user)
def uploaded_files_view(request):
    files = UploadedFile.objects.select_related('user', 'custom_tag').order_by('-timestamp')
    return render(request, 'capture/uploaded_files_table.html', {'files': files})

@login_required
@user_passes_test(is_super_user)
def uploaded_images_view(request):
    files = UploadedImage.objects.select_related('user', 'custom_tag').order_by('-timestamp')
    return render(request, 'capture/uploaded_images_table.html', {'files': files})

@login_required
@user_passes_test(is_super_user)
def captured_images_view(request):
    files = CapturedImage.objects.select_related('user', 'custom_tag').order_by('-timestamp')
    return render(request, 'capture/captured_images_table.html', {'files': files})


@login_required
@user_passes_test(is_super_user)
def all_files_images_view(request):
    # Fetch data from all three models
    uploaded_files = UploadedFile.objects.select_related('user', 'custom_tag').order_by('-timestamp')
    uploaded_images = UploadedImage.objects.select_related('user', 'custom_tag').order_by('-timestamp')
    captured_images = CapturedImage.objects.select_related('user', 'custom_tag').order_by('-timestamp')

    # Add source_type attribute to each item dynamically
    for obj in uploaded_files:
        obj.source_type = "Uploaded File"
    for obj in uploaded_images:
        obj.source_type = "Uploaded Image"
    for obj in captured_images:
        obj.source_type = "Captured Image"

    # Merge all into a single list and sort by timestamp
    all_files = sorted(
        chain(uploaded_files, uploaded_images, captured_images),
        key=lambda x: x.timestamp,
        reverse=True
    )

    return render(request, 'capture/all_files.html', {'files': all_files})



@login_required
@user_passes_test(is_super_user)
def export_uploaded_files_excel(request):
    files = UploadedFile.objects.select_related('user', 'custom_tag').order_by('-timestamp')

    # Create workbook and worksheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Uploaded Files"

    # Add header row with updated columns
    headers = ['Sl.', 'IP', 'Tag', 'Name', 'Username', 'Designation', 'Department', 'Phone', 'Date-Time','File Type', 'File Size (MB)', 'File Link']
    ws.append(headers)

    # Get the base URL for the files (e.g., "http://127.0.0.1:8000")
    base_url = settings.SITE_URL if hasattr(settings, 'SITE_URL') else request.build_absolute_uri('/').split('/')[0] + '//' + request.get_host()

    def bytes_to_mb(size_bytes):
        if not size_bytes:
            return 'N/A'
        return f"{size_bytes / (1024 * 1024):.2f} MB"
    
    # Add data rows with updated order
    for idx, file in enumerate(files, start=1):
        designation = file.user.designation.title if file.user.designation else 'N/A'
        if file.user.designation_other:
            designation += f" ({file.user.designation_other})"
        
        department = file.user.department.name if file.user.department else 'N/A'
        if file.user.department_other:
            department += f" ({file.user.department_other})"
        
        # Format UHID like 25/1678
        uhid_str = str(file.uhid)
        formatted_uhid = f"{uhid_str[:2]}/{uhid_str[2:]}" if len(uhid_str) > 2 else uhid_str
        
        # Construct the full file URL
        file_url = f'{base_url}{file.file_path.url}' if file.file_path else ''

        # Convert timestamp to local time and format
        local_timestamp = timezone.localtime(file.timestamp)
        datetime_str = local_timestamp.strftime('%d-%b-%Y %I:%M %p')

        # Get file type (from model field if available)
        file_type = file.file_type if file.file_type else 'Unknown'

        # Get file size in MB
        file_size_str = bytes_to_mb(file.file_size)



        # Adding data in the specified order
        ws.append([
            idx,  # Sl.
            formatted_uhid,  # IP (using the `uhid` for IP as per your data)
            file.custom_tag.name if file.custom_tag else '—',  # Tag (showing the custom_tag or '—')
            file.user.full_name,  # Name
            file.user.email,  # Username
            designation,  # Designation (including designation_other if exists)
            department,  # Department (including department_other if exists)
            file.user.phone if file.user.phone else 'N/A',
            datetime_str,
            file_type,  # File Type
            file_size_str,  # File Size (MB)
            f'=HYPERLINK("{file_url}", "Click Here")' if file_url else 'No Link',  # File Link (full URL for the file)
        ])

    # Generate dynamic filename with date-time
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f"uploaded_files_{timestamp}.xlsx"

    # Prepare HTTP response with Excel content
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename={filename}'
    wb.save(response)

    return response



@login_required
@user_passes_test(is_super_user)
def export_uploaded_images_excel(request):
    files = UploadedImage.objects.select_related('user', 'custom_tag').order_by('-timestamp')

    # Create workbook and worksheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Uploaded Images"

    # Add header row
    headers = ['Sl.', 'IP', 'Tag', 'Name', 'Username', 'Designation', 'Department', 'Phone', 'Date-Time','File Type', 'File Size (MB)', 'File Link']
    ws.append(headers)

    # Get the base URL for the files (e.g., "http://127.0.0.1:8000")
    base_url = settings.SITE_URL if hasattr(settings, 'SITE_URL') else request.build_absolute_uri('/').split('/')[0] + '//' + request.get_host()

    def bytes_to_mb(size_bytes):
        if not size_bytes:
            return 'N/A'
        return f"{size_bytes / (1024 * 1024):.2f} MB"

    def get_file_type(file_obj):
        if file_obj.image_path:
            ext = file_obj.image_path.name.split('.')[-1].lower()
            return f'image/{ext}'
        return 'Unknown'
    
    # Add data rows with updated order
    for idx, file in enumerate(files, start=1):
        designation = file.user.designation.title if file.user.designation else 'N/A'
        if file.user.designation_other:
            designation += f" ({file.user.designation_other})"
        
        department = file.user.department.name if file.user.department else 'N/A'
        if file.user.department_other:
            department += f" ({file.user.department_other})"
        
        # Format UHID like 25/1678
        uhid_str = str(file.uhid)
        formatted_uhid = f"{uhid_str[:2]}/{uhid_str[2:]}" if len(uhid_str) > 2 else uhid_str
        
        # Construct the full file URL
        file_url = f'{base_url}{file.image_path.url}' if file.image_path else ''

        # Convert timestamp to local time and format
        local_timestamp = timezone.localtime(file.timestamp)
        datetime_str = local_timestamp.strftime('%d-%b-%Y %I:%M %p')

        # Get file type and size
        file_type = get_file_type(file)
        file_size_str = bytes_to_mb(file.image_size)

        # Adding data in the specified order
        ws.append([
            idx,  # Sl.
            formatted_uhid,  # IP (using the `uhid` for IP as per your data)
            file.custom_tag.name if file.custom_tag else '—',  # Tag (showing the custom_tag or '—')
            file.user.full_name,  # Name
            file.user.email,  # Username
            designation,  # Designation (including designation_other if exists)
            department,  # Department (including department_other if exists)
            file.user.phone if file.user.phone else 'N/A',
            datetime_str,
            file_type,  # File Type
            file_size_str,  # File Size (MB)
            f'=HYPERLINK("{file_url}", "Click Here")' if file_url else 'No Link',  # File Link (full URL for the file)
        ])

    # Generate dynamic filename with date-time
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f"uploaded_images_{timestamp}.xlsx"

    # Prepare HTTP response with Excel content
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename={filename}'
    wb.save(response)

    return response

@login_required
@user_passes_test(is_super_user)
def export_captured_images_excel(request):
    files =CapturedImage.objects.select_related('user', 'custom_tag').order_by('-timestamp')

    # Create workbook and worksheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Captured Images"

    # Add header row
    headers = ['Sl.', 'IP', 'Tag', 'Name', 'Username', 'Designation', 'Department', 'Phone', 'Date-Time','File Type', 'File Size (MB)', 'File Link']
    ws.append(headers)

    # Get the base URL for the files (e.g., "http://127.0.0.1:8000")
    base_url = settings.SITE_URL if hasattr(settings, 'SITE_URL') else request.build_absolute_uri('/').split('/')[0] + '//' + request.get_host()

    def bytes_to_mb(size_bytes):
        if not size_bytes:
            return 'N/A'
        return f"{size_bytes / (1024 * 1024):.2f} MB"

    def get_file_type(file_obj):
        # Guess file type from image_path extension
        if file_obj.image_path:
            ext = file_obj.image_path.name.split('.')[-1].lower()
            return f'image/{ext}'
        return 'Unknown'
    
    # Add data rows with updated order
    for idx, file in enumerate(files, start=1):
        designation = file.user.designation.title if file.user.designation else 'N/A'
        if file.user.designation_other:
            designation += f" ({file.user.designation_other})"
        
        department = file.user.department.name if file.user.department else 'N/A'
        if file.user.department_other:
            department += f" ({file.user.department_other})"
        
        # Format UHID like 25/1678
        uhid_str = str(file.uhid)
        formatted_uhid = f"{uhid_str[:2]}/{uhid_str[2:]}" if len(uhid_str) > 2 else uhid_str
        
        # Construct the full file URL
        file_url = f'{base_url}{file.image_path.url}' if file.image_path else ''

        # Convert timestamp to local time and format as '12-May-2025 10:15 AM'
        local_timestamp = timezone.localtime(file.timestamp)
        datetime_str = local_timestamp.strftime('%d-%b-%Y %I:%M %p')
        
        # Get file type and size
        file_type = get_file_type(file)
        file_size_str = bytes_to_mb(file.image_size)

        # Adding data in the specified order
        ws.append([
            idx,  # Sl.
            formatted_uhid,  # IP (using the `uhid` for IP as per your data)
            file.custom_tag.name if file.custom_tag else '—',  # Tag (showing the custom_tag or '—')
            file.user.full_name,  # Name
            file.user.email,  # Username
            designation,  # Designation (including designation_other if exists)
            department,  # Department (including department_other if exists)
            file.user.phone if file.user.phone else 'N/A',
            datetime_str,
            file_type,  # File Type
            file_size_str,  # File Size (MB)
            f'=HYPERLINK("{file_url}", "Click Here")' if file_url else 'No Link',  # File Link (full URL for the file)
        ])


    # Generate dynamic filename with date-time
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f"captured_images_{timestamp}.xlsx"

    # Prepare HTTP response with Excel content
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename={filename}'
    wb.save(response)

    return response

@login_required
@user_passes_test(is_super_user)
def export_all_files_excel(request):
    # Fetch records from all 3 models
    uploaded_files = UploadedFile.objects.select_related('user', 'custom_tag').order_by('-timestamp')
    uploaded_images = UploadedImage.objects.select_related('user', 'custom_tag').order_by('-timestamp')
    captured_images = CapturedImage.objects.select_related('user', 'custom_tag').order_by('-timestamp')

    # Create workbook and worksheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "All Files & Images"

    # Header row
    headers = ['Sl.', 'Type', 'IP', 'Tag', 'Name', 'Username', 'Designation', 'Department', 'Phone', 'Date-Time','File Type', 'File Size (MB)', 'File Link']
    ws.append(headers)

    # Base URL
    base_url = settings.SITE_URL if hasattr(settings, 'SITE_URL') else request.build_absolute_uri('/').split('/')[0] + '//' + request.get_host()

    def bytes_to_mb(size_bytes):
        if not size_bytes:
            return 'N/A'
        return f"{size_bytes / (1024 * 1024):.2f} MB"
    
    def get_file_type(file_obj, source_type):
        # For UploadedFile, file_type field exists
        if source_type == 'Uploaded File' and file_obj.file_type:
            return file_obj.file_type
        # For images, guess from extension
        if source_type in ['Uploaded Image', 'Captured Image']:
            if hasattr(file_obj, 'image_path') and file_obj.image_path:
                ext = file_obj.image_path.name.split('.')[-1].lower()
                return f'image/{ext}'
        return 'Unknown'

    def get_file_size(file_obj, source_type):
        # UploadedFile has file_size in bytes
        if source_type == 'Uploaded File' and file_obj.file_size:
            return file_obj.file_size
        # UploadedImage and CapturedImage have image_size field
        if source_type in ['Uploaded Image', 'Captured Image']:
            if hasattr(file_obj, 'image_size') and file_obj.image_size:
                return file_obj.image_size
        return None    

    def write_row(idx, file, source_type, file_url):
        # Format designation
        designation = file.user.designation.title if file.user.designation else 'N/A'
        if file.user.designation_other:
            designation += f" ({file.user.designation_other})"
        
        # Format department
        department = file.user.department.name if file.user.department else 'N/A'
        if file.user.department_other:
            department += f" ({file.user.department_other})"
        
        # Format UHID
        uhid_str = str(file.uhid)
        formatted_uhid = f"{uhid_str[:2]}/{uhid_str[2:]}" if len(uhid_str) > 2 else uhid_str

        # Convert timestamp to local time before formatting
        local_timestamp = timezone.localtime(file.timestamp)
        datetime_str = local_timestamp.strftime('%d-%b-%Y %I:%M %p')

        # File type and size
        file_type = get_file_type(file, source_type)
        file_size_bytes = get_file_size(file, source_type)
        file_size_str = bytes_to_mb(file_size_bytes)

        # Add row
        ws.append([
            idx,
            source_type,
            formatted_uhid,
            file.custom_tag.name if file.custom_tag else '—',
            file.user.full_name,
            file.user.email,
            designation,
            department,
            file.user.phone if file.user.phone else 'N/A',
            datetime_str,
            file_type,
            file_size_str,
            f'=HYPERLINK("{file_url}", "Click Here")' if file_url else 'No Link',
        ])

    # Combined list with labels
    all_files = [
        ('Uploaded File', uploaded_files, lambda f: f.file_path.url if f.file_path else ''),
        ('Uploaded Image', uploaded_images, lambda f: f.image_path.url if f.image_path else ''),
        ('Captured Image', captured_images, lambda f: f.image_path.url if f.image_path else ''),
    ]

    idx = 1
    for label, queryset, path_getter in all_files:
        for file in queryset:
            file_url = f'{base_url}{path_getter(file)}' if path_getter(file) else ''
            write_row(idx, file, label, file_url)
            idx += 1

    # Generate filename
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f"all_files_and_images_{timestamp}.xlsx"

    # Return Excel response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename={filename}'
    wb.save(response)
    return response

@login_required
def report_issue(request):
    print(f"[DEBUG] User: {request.user.email} - {request.user.get_full_name()}")

    # 1. Count user's pending issues (status ≠ 'completed')
    pending_count = (
        IssueReport.objects
        .filter(user=request.user, is_deleted=False)
        .exclude(current_status="completed")
        .count()
    )
    form_allowed = pending_count < 3
    print(f"[DEBUG] Pending issues count: {pending_count}, Form allowed: {form_allowed}")

    if request.method == "POST":
        print("[DEBUG] POST request received")
        if not form_allowed:
            print("[DEBUG] Form submission blocked due to issue limit.")
            messages.error(
                request,
                "Maximum pending issues reached. Please wait until one is completed.",
                extra_tags='no-dismiss'
            )
            return redirect("capture:report_issue")

        form = IssueReportForm(request.POST, request.FILES)
        if form.is_valid():
            issue = form.save(commit=False)
            issue.user = request.user
            issue.save()
            print(f"[DEBUG] Issue saved: ID = {issue.issue_id}")

            # Email to admin
            subject = f"[New Issue] {issue.issue_id} from {request.user.get_full_name() or request.user.email}"
            body = (
                f"A new issue has been reported.\n\n"
                f"User: {request.user.get_full_name() or request.user.email}\n"
                f"Issue ID: {issue.issue_id}\n"
                f"Status: {issue.get_current_status_display()}\n"
                f"Description:\n{issue.description}\n\n"
                f"Submitted at: {issue.created_at.strftime('%Y-%m-%d %H:%M:%S')}\n"
            )
            try:
                send_mail(
                    subject,
                    body,
                    settings.DEFAULT_FROM_EMAIL,
                    [settings.ADMIN_EMAIL],
                    fail_silently=False,
                )
                print("[DEBUG] Admin email sent successfully.")
            except Exception as e:
                print(f"[DEBUG] Email sending failed: {e}")

            messages.success(request, "Your issue has been reported successfully.")
            return redirect("capture:report_issue")
        else:
            print(f"[DEBUG] Form errors: {form.errors}")
            messages.error(request, "Please correct the errors below.")
    else:
        print("[DEBUG] GET request received")
        if not form_allowed:
            messages.error(
                request,
                "You have reached the maximum number of pending issues (3). Please wait until one is completed.",
                extra_tags='no-dismiss'
            )
        form = IssueReportForm() if form_allowed else None

    # 2. Get user issues for display
    user_issues = (
        IssueReport.objects
        .filter(user=request.user, is_deleted=False)
        .order_by("-created_at")
    )
    print(f"[DEBUG] Total issues retrieved for table: {user_issues.count()}")

    return render(
        request,
        "capture/report_issue.html",
        {
            "form": form,
            "user_issues": user_issues,
            "form_allowed": form_allowed
        },
    )


@login_required
@user_passes_test(is_super_user)
def manage_issues(request):
    if request.method == "POST":
        issue = get_object_or_404(IssueReport, pk=request.POST.get("issue_id"))
        form = IssueAdminUpdateForm(request.POST, instance=issue)
        if form.is_valid():
            issue = form.save(commit=False)
            issue.current_status_marked_at = timezone.now()
            issue.save()
            messages.success(request, f"Issue {issue.issue_id} updated successfully.")
        else:
            messages.error(request, "There was an error updating the issue.")
        return redirect("capture:manage_issues")

    all_issues = IssueReport.objects.filter(is_deleted=False).order_by("-created_at")
    issue_forms = [IssueAdminUpdateForm(instance=issue) for issue in all_issues]
    issue_data = zip(all_issues, issue_forms)

    return render(
        request,
        "capture/manage_issues.html",
        {
            "issue_data": issue_data,  # List of (issue, form) pairs
        },
    )











