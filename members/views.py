from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import login
from django.contrib.auth.decorators import login_required, user_passes_test
from .forms import UserRegisterForm, UserLoginForm, ProfileForm, CustomPasswordChangeForm, CustomTagForm, DepartmentForm, DesignationForm, UploadExcelForm, WardForm
from django.contrib import messages
from django.contrib.auth import update_session_auth_hash
from django.contrib.auth.forms import PasswordChangeForm
from django.contrib.auth.forms import AuthenticationForm
from django.contrib.auth import authenticate
from django.views.decorators.csrf import csrf_exempt
from django.http import JsonResponse
import json
from .models import CustomUser, CustomTag, Department, Designation, Ward
from django.utils import timezone
from django.db.models import Q
from django.contrib.auth.hashers import make_password
from django.core.mail import send_mail, BadHeaderError
from django.conf import settings
import logging
from django.contrib.auth import get_user_model
from django.core.exceptions import ObjectDoesNotExist
from django.http import HttpResponse
from django.utils.timezone import localtime
from openpyxl import Workbook
from datetime import datetime
import csv
import io
import openpyxl
from django.views.decorators.cache import never_cache
from capture.models import UploadedFile, UploadedImage, CapturedImage
from django.db.models import Count
from django.db.models import Count, Q, F, Value, IntegerField, Sum


# Get the logger for emails
logger = logging.getLogger('django.core.mail')


# 1️⃣ User Registration View
@never_cache
def register(request):
    if request.method == "POST":
        form = UserRegisterForm(request.POST)
        print("POST raw department_other:", request.POST.get('department_other'))
        print("POST raw designation_other:", request.POST.get('designation_other'))
        print("POST raw ward_other:", request.POST.get('ward_other'))
        
        if form.is_valid():
            print("REGISTER cleaned_data:", form.cleaned_data)
            user = form.save()
            print(f"✅ SUCCESS: User {user.email} registered successfully!")

            # Prepare Welcome Email
            subject = 'Welcome to MedCap'
            message = (
                f"Hi {user.full_name},\n\n"
                "Welcome to MedCap!\n\n"
                "Thank you for registering. We're excited to have you on board.\n\n"
                "If you have any questions, feel free to reach out to us at medcap.srhu@gmail.com.\n\n"
                "Best regards,\n"
                "MedCap Team- SRHU"
            )
            from_email = settings.DEFAULT_FROM_EMAIL
            recipient_list = [user.email]

            try:
                send_mail(
                    subject,
                    message,
                    from_email,
                    recipient_list,
                    fail_silently=False  # We WANT to catch the error here
                )
                logger.debug(f"✅ Welcome email sent successfully to {user.email}")
            except BadHeaderError as e:
                logger.error(f"❌ BadHeaderError while sending email to {user.email}: {e}")
            except Exception as e:
                logger.error(f"❌ Unexpected error while sending welcome email to {user.email}: {e}")

            messages.success(
                request,
                "✅ Registration successful. Account sent for Approval."
            )
            return redirect("members:login")
    else:
        form = UserRegisterForm()

    return render(request, "members/register.html", {"form": form})

def is_profile_complete(user):
    # Check all required fields are filled and not empty
    required_fields = [
        user.full_name,
        user.email,
        user.department,
        user.designation,
        user.phone
    ]
    return all(required_fields)


@never_cache
def login_view(request):
    if request.method == "POST":
        form = UserLoginForm(request, data=request.POST)
        email = request.POST.get('username').strip().lower()  # 'username' field holds email in Django auth forms
        password = request.POST.get('password')

        try:
            user = get_user_model().objects.get(email=email)
            print(f"DEBUG: Found user with email: {email}")
        except ObjectDoesNotExist:
            user = None
            print(f"DEBUG: No user found with email: {email}")

        if form.is_valid():
            # Form is valid: credentials matched

            # Debugging: Check if the user is approved
            print(f"DEBUG: Attempting to login user with email: {user.email}")
            print(f"DEBUG: Initial user approval status: {user.is_approved}")
            print(f"DEBUG: Initial failed attempts count: {user.failed_attempts}")

            # Check if the user is a superuser or approved
            if user.is_superuser:
                print(f"DEBUG: User {user.email} is a superuser, bypassing approval check.")
            elif not user.is_approved:
                print(f"DEBUG: User {user.email} is not approved. Approval pending.")
                messages.error(request, "Approval pending")
                return redirect('members:login')

            # Handle failed attempts and account locking logic
            if user.failed_attempts >= 6:
                print(f"DEBUG: User {user.email} has {user.failed_attempts} failed attempts, account locked.")
                messages.error(request, "Your account is locked due to too many failed attempts. Please contact support.")
                return redirect('members:login')

            if user.failed_attempts >= 4:
                print(f"DEBUG: User {user.email} has {user.failed_attempts} failed attempts, warning issued.")
                messages.warning(request, "You have entered wrong password 4 times, 2 more failed attempts will block the account. Please Reset your password if you have forgotten.")

            # Successful authentication
            print(f"DEBUG: User {user.email} successfully authenticated.")
            user.failed_attempts = 0  # Reset failed attempts on success
            user.save()

            login(request, user)

            if user.is_superuser:
                print(f"DEBUG: Superuser {user.email} redirected to configurations page.")
                return redirect('members:configs')
            
            # Check profile completeness here
            if not is_profile_complete(user):
                messages.info(request, "Please complete your profile before using the app.")
                return redirect('members:profile_update')  # Your profile update page URL name

            print(f"DEBUG: Regular user {user.email} redirected to capture page.")
            return redirect('members:service_choice')

        else:
            # Form invalid: bad credentials
            print(f"DEBUG: Form is invalid. Errors: {form.errors}")

            if user:
                # User exists, so password must be wrong
                print(f"DEBUG: Invalid password entered for user {user.email}. Incrementing failed attempts.")
                user.failed_attempts += 1
                user.save()

                if user.failed_attempts >= 6:
                    if user.is_active:  # Check if the account is active and only lock it once
                        print(f"DEBUG: User {user.email} reached 6 failed attempts, locking account.")
                        user.is_active = False
                        user.save()

                        # Send lock email only when the user is blocked for the first time
                        try:
                            send_mail(
                                'Account Locked Due to Failed Login Attempts',
                                f'User {user.email} has been locked out after 6 failed login attempts.',
                                settings.DEFAULT_FROM_EMAIL,
                                [settings.ADMIN_EMAIL],
                            )
                        except Exception as e:
                            print(f"DEBUG: Failed to send lock email for user {user.email}: {e}")

                        messages.error(request, "Your account has been locked. Please contact support.")
                    else:
                        # User is already locked, no need to lock again, just show the message
                        messages.error(request, "Your account had been locked. Please contact support.")

                else:
                    if user.failed_attempts >= 4:
                        print(f"DEBUG: Warning user {user.email} for multiple failed attempts.")
                        messages.warning(request, "You have entered wrong password multiple times. Please reset your password from Forget Password link.")
                    messages.error(request, "Invalid login credentials. Please try again.")

            else:
                # No such user
                print(f"DEBUG: No such user with email: {email}")
                messages.error(request, "User not found. Please check your email.")

            return redirect('members:login')

    else:
        form = UserLoginForm()

    return render(request, "members/login.html", {"form": form})


#to select the type of service ayushman or echs
def service_choice(request):
    return render(request, 'members/service_choice.html')

# Only allow superusers to approve other users
def is_super_user(user):
    return user.is_superuser

@user_passes_test(is_super_user)
@login_required
def user_approval(request):
    """
    View to handle user approval by Superuser.
    """
    all_users = CustomUser.objects.filter(~Q(is_superuser=True)).order_by('full_name')

    if request.method == "POST":
        user_id = request.POST.get("user_id")  # Get user ID from the POST data
        action = request.POST.get("action")  # 'approve' or 'disapprove'

        try:
            user = CustomUser.objects.get(id=user_id)
            
            if action == "approve":
                user.is_approved = True
                user.approved_by = request.user
                user.approved_at = timezone.now()
                user.save()

                # ✅ Send Account Approval Email
                subject = 'Your MedCap Account is Approved'
                message = (
                    f"Hi {user.full_name},\n\n"
                    "Good news! Your MedCap account has been approved and is now active.\n\n"
                    "You can now log in and start using all the features.\n\n"
                    "If you have any questions, feel free to reach out at medcap.srhu@gmail.com.\n\n"
                    "Best regards,\n"
                    "MedCap Team"
                )
                from_email = settings.DEFAULT_FROM_EMAIL
                recipient_list = [user.email]

                try:
                    send_mail(
                        subject,
                        message,
                        from_email,
                        recipient_list,
                        fail_silently=False
                    )
                    logger.debug(f"✅ Approval email sent successfully to {user.email}")
                except BadHeaderError as e:
                    logger.error(f"❌ BadHeaderError while sending approval email to {user.email}: {e}")
                except Exception as e:
                    logger.error(f"❌ Unexpected error while sending approval email to {user.email}: {e}")

                messages.success(request, f"User {user.email} has been approved.")
                
            elif action == "disapprove":
                user.is_approved = False
                user.approved_by = None  # Optionally clear approved_by if disapproved
                user.approved_at = None  # Optionally clear approved_at if disapproved
                user.save()
                messages.info(request, f"User {user.email} has been disapproved.")
            else:
                messages.error(request, "Invalid action.")

        except CustomUser.DoesNotExist:
            messages.error(request, "User not found.")

        return redirect('members:user_approval')  # Redirect back to approval page after action

    return render(request, 'members/user_approval.html', {
        'all_users': all_users
    })

@login_required
@user_passes_test(is_super_user)
def reset_password_bysuper(request):
    """
    Reset the password for a specific user to '123456'.
    """
    if request.method == 'POST':
        user_id = request.POST.get('user_id')
        try:
            # Fetch the user from the database
            user = CustomUser.objects.get(id=user_id)
            
            # Set the password to '123456'
            user.password = make_password('123456')
            user.save()

            # Show success message
            messages.success(request, f"Password for {user.email} has been reset to '123456'.")
        except CustomUser.DoesNotExist:
            messages.error(request, "User not found.")

    return redirect('members:user_approval')

@user_passes_test(is_super_user)
@login_required
def unblock_user_bysuper(request):
    """
    Unblock a user who was blocked due to failed login attempts.
    Sets is_active to True and resets failed_attempts to 0.
    """
    if request.method == "POST":
        user_id = request.POST.get("user_id")

        try:
            # Fetch the user from the database
            user = CustomUser.objects.get(id=user_id)

            # Perform the unblock action
            user.is_active = True
            user.failed_attempts = 0
            user.save()

            messages.success(request, f"User {user.email} has been unblocked successfully.")

        except CustomUser.DoesNotExist:
            messages.error(request, "User not found.")

    return redirect('members:user_approval')


@user_passes_test(is_super_user)
@login_required
def configs(request):
    # Cache all departments and users once
    departments = Department.objects.all()
    active_department_ids = list(departments.values_list('id', flat=True))

    all_users = CustomUser.objects.select_related('department').all()

    # Pre-fetch file/image counts per user efficiently by aggregating
    # We'll do a single query per model grouped by user, then map later

    # Aggregated counts per user for UploadedFile
    uploaded_file_counts = UploadedFile.objects.values('user').annotate(count=Count('id'))
    uploaded_file_map = {item['user']: item['count'] for item in uploaded_file_counts}

    # Aggregated counts per user for CapturedImage
    captured_image_counts = CapturedImage.objects.values('user').annotate(count=Count('id'))
    captured_image_map = {item['user']: item['count'] for item in captured_image_counts}

    # Aggregated counts per user for UploadedImage
    uploaded_image_counts = UploadedImage.objects.values('user').annotate(count=Count('id'))
    uploaded_image_map = {item['user']: item['count'] for item in uploaded_image_counts}

    # Calculate user_stats efficiently without per-user queries
    user_stats = []
    for user in all_users:
        uf_count = uploaded_file_map.get(user.id, 0)
        ci_count = captured_image_map.get(user.id, 0)
        ui_count = uploaded_image_map.get(user.id, 0)
        total_files = uf_count + ci_count + ui_count

        user_stats.append({
            'name': user.get_full_name() or user.username,
            'total_files': total_files,
        })

    # Department-wise stats using ORM annotation (to minimize queries)
    # First, get user ids per department (cached)
    dept_user_map = {}
    for dept in departments:
        dept_user_map[dept.id] = list(all_users.filter(department=dept).values_list('id', flat=True))

    department_stats = []
    sum_users = 0
    sum_images = 0
    sum_files = 0
    sum_total = 0

    for i, dept in enumerate(departments, start=1):
        user_ids = dept_user_map.get(dept.id, [])

        # Count files/images for users in this department using pre-aggregated counts
        captured_count = CapturedImage.objects.filter(user_id__in=user_ids).count()
        uploaded_image_count = UploadedImage.objects.filter(user_id__in=user_ids).count()
        uploaded_file_count = UploadedFile.objects.filter(user_id__in=user_ids).count()
        total_image_count = captured_count + uploaded_image_count
        total = total_image_count + uploaded_file_count

        user_count = len(user_ids)
        sum_users += user_count
        sum_images += total_image_count
        sum_files += uploaded_file_count
        sum_total += total

        department_stats.append({
            'sno': i,
            'name': dept.name,
            'user_count': user_count,
            'image_count': total_image_count,
            'captured_count': captured_count,
            'uploaded_image_count': uploaded_image_count,
            'file_count': uploaded_file_count,
            'total': total,
        })

    # Overall counts computed once from cached querysets where possible
    total_departments = departments.count()
    total_users = all_users.count()

    # For totals, use aggregated counts instead of multiple queries
    total_uploaded_files = UploadedFile.objects.count()
    total_captured_images = CapturedImage.objects.count()
    total_uploaded_images = UploadedImage.objects.count()

    total_image_files = total_captured_images + total_uploaded_images
    total_files = total_uploaded_files + total_image_files

    total_tags = CustomTag.objects.filter(is_deleted=False).count()
    tags = CustomTag.objects.filter(is_deleted=False)

    # File/Image Querysets for Modals, ordered descending by timestamp
    uploaded_files = UploadedFile.objects.order_by('-timestamp')
    captured_images = CapturedImage.objects.order_by('-timestamp')
    uploaded_images = UploadedImage.objects.order_by('-timestamp')

    context = {
        'total_departments': total_departments,
        'total_users': total_users,
        'total_files': total_files,
        'total_uploaded_files': total_uploaded_files,
        'total_captured_images': total_captured_images,
        'total_uploaded_images': total_uploaded_images,
        'total_image_files': total_image_files,
        'total_tags': total_tags,
        'uploaded_files': uploaded_files,
        'captured_images': captured_images,
        'uploaded_images': uploaded_images,
        'department_stats': department_stats,
        'sum_users': sum_users,
        'sum_images': sum_images,
        'sum_files': sum_files,
        'sum_total': sum_total,
        'user_stats': user_stats,
        'tags': tags,
    }

    return render(request, 'members/configs.html', context)


# Superuser-only access
@user_passes_test(lambda u: u.is_superuser)
def manage_custom_tags(request):
    # Handle the form for adding new custom tags
    if request.method == 'POST':
        form = CustomTagForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Added successfully!')
            return redirect('members:manage_custom_tags')
    else:
        form = CustomTagForm()

    # Separate active and inactive tags
    active_tags = CustomTag.objects.filter(is_deleted=False).order_by('name')
    inactive_tags = CustomTag.objects.filter(is_deleted=True).order_by('name')

    return render(request, 'members/manage_custom_tags.html', {
        'form': form,
        'active_tags': active_tags,
        'inactive_tags': inactive_tags
    })

@user_passes_test(lambda u: u.is_superuser)
def delete_custom_tag(request, pk):
    tag = get_object_or_404(CustomTag, pk=pk)
    tag.is_deleted = True
    tag.save()
    messages.success(request, f"Tag inactivated")
    return redirect('members:manage_custom_tags')

@user_passes_test(lambda u: u.is_superuser)
def restore_custom_tag(request, pk):
    tag = get_object_or_404(CustomTag, pk=pk)
    tag.is_deleted = False
    tag.save()
    messages.success(request, f"Tag '{tag.name}' has been restored.")
    return redirect('members:manage_custom_tags')


@user_passes_test(lambda u: u.is_superuser)
def manage_departments(request):
    if request.method == 'POST':
        form = DepartmentForm(request.POST)
        if form.is_valid():
            dept = form.save(commit=False)
            dept.created_by = request.user
            dept.save()
            messages.success(request, 'Department added successfully!')
            return redirect('members:manage_departments')
    else:
        form = DepartmentForm()

    active_departments = Department.objects.filter(is_active=True).order_by('name')
    inactive_departments = Department.objects.filter(is_active=False).order_by('name')

    return render(request, 'members/manage_departments.html', {
        'form': form,
        'active_departments': active_departments,
        'inactive_departments': inactive_departments,
    })


@user_passes_test(lambda u: u.is_superuser)
def deactivate_department(request, pk):
    dept = get_object_or_404(Department, pk=pk)
    dept.is_active = False
    dept.save()
    messages.success(request, f"Department '{dept.name}' deactivated.")
    return redirect('members:manage_departments')

@user_passes_test(lambda u: u.is_superuser)
def restore_department(request, pk):
    dept = get_object_or_404(Department, pk=pk)
    dept.is_active = True
    dept.save()
    messages.success(request, f"Department '{dept.name}' restored.")
    return redirect('members:manage_departments')

@user_passes_test(lambda u: u.is_superuser)
def manage_wards(request):
    if request.method == 'POST':
        form = WardForm(request.POST)
        if form.is_valid():
            ward = form.save(commit=False)
            ward.created_by = request.user
            ward.save()
            messages.success(request, 'Ward added successfully!')
            return redirect('members:manage_wards')
    else:
        form = WardForm()

    active_wards = Ward.objects.filter(is_active=True).order_by('ward_name')
    inactive_wards = Ward.objects.filter(is_active=False).order_by('ward_name')

    return render(request, 'members/manage_wards.html', {
        'form': form,
        'active_wards': active_wards,
        'inactive_wards': inactive_wards,
    })

@user_passes_test(lambda u: u.is_superuser)
def deactivate_ward(request, pk):
    ward = get_object_or_404(Ward, pk=pk)
    ward.is_active = False
    ward.save()
    messages.success(request, f"Ward '{ward.ward_name}' deactivated.")
    return redirect('members:manage_wards')

@user_passes_test(lambda u: u.is_superuser)
def restore_ward(request, pk):
    ward = get_object_or_404(Ward, pk=pk)
    ward.is_active = True
    ward.save()
    messages.success(request, f"Ward '{ward.ward_name}' restored.")
    return redirect('members:manage_wards')


@user_passes_test(lambda u: u.is_superuser)
def manage_designations(request):
    if request.method == 'POST':
        form = DesignationForm(request.POST)
        if form.is_valid():
            designation = form.save(commit=False)
            designation.created_by = request.user
            designation.save()
            messages.success(request, 'Designation added successfully!')
            return redirect('members:manage_designations')
    else:
        form = DesignationForm()

    active_designations = Designation.objects.filter(is_active=True).order_by('title')
    inactive_designations = Designation.objects.filter(is_active=False).order_by('title')

    return render(request, 'members/manage_designations.html', {
        'form': form,
        'active_designations': active_designations,
        'inactive_designations': inactive_designations,
    })

@user_passes_test(lambda u: u.is_superuser)
def deactivate_designation(request, pk):
    d = get_object_or_404(Designation, pk=pk)
    d.is_active = False
    d.save()
    messages.success(request, f"Designation '{d.title}' deactivated.")
    return redirect('members:manage_designations')

@user_passes_test(lambda u: u.is_superuser)
def restore_designation(request, pk):
    d = get_object_or_404(Designation, pk=pk)
    d.is_active = True
    d.save()
    messages.success(request, f"Designation '{d.title}' restored.")
    return redirect('members:manage_designations')



@login_required
def profile_update(request):
 
    # Debugging: Print request method (POST or GET)
    print(f"Request Method: {request.method}")

    if request.method == 'POST':
        # Handling Profile Form submission
        profile_form = ProfileForm(request.POST, instance=request.user)
        
        # Debugging: Print form data
        print(f"Profile Form Data: {profile_form.cleaned_data if profile_form.is_valid() else 'Invalid'}")

        # If both forms are valid, save the data and update the session
        if profile_form.is_valid():
            # Debugging: Notify when forms are valid and data will be saved
            print("Profile Form is valid. Saving data...")

            # Save Profile Form data
            profile_form.save()

            # Show success message
            messages.success(request, "Your profile updated successfully!")
            return redirect('members:profile')  # Redirect to profile page
        else:
            # Debugging: Print form errors if any
            print(f"Form Errors: Profile Form: {profile_form.errors}")

            # If forms are invalid, show error message
            messages.error(request, "Please correct the errors below.")
    
    else:
        # GET request: Pre-fill the forms with the current user’s data
        profile_form = ProfileForm(instance=request.user)

        # Debugging: Notify that we are rendering the page with pre-filled forms
        print("Rendering page with pre-filled forms.")

    # Render the profile update template with the forms
    return render(request, 'members/profile_update.html', {
        'profile_form': profile_form
    })

@login_required
def password_update(request):

    # Debugging: Print request method (POST or GET)
    print(f"Request Method: {request.method}")

    if request.method == 'POST':
        # Handling Password Form submission
        password_form = CustomPasswordChangeForm(request.user, request.POST)
        
        # Debugging: Print form data
        print(f"Password Form Data: {password_form.cleaned_data if password_form.is_valid() else 'Invalid'}")

        # If Form is valid, save the data and update the session
        if password_form.is_valid():
            # Debugging: Notify when forms are valid and data will be saved
            print("Password Forms is valid. Saving data...")

            # Save password change data
            password_form.save()
            
            # Update session to prevent user from getting logged out after password change
            update_session_auth_hash(request, password_form.user)

            # Show success message
            messages.success(request, "Your password updated successfully!")
            return redirect('members:profile')  # Redirect to profile page
        else:
            # Debugging: Print form errors if any
            print(f"Form Errors: Password Form: {password_form.errors}")

            # If forms are invalid, show error message
            messages.error(request, "Please correct the errors below.")
    
    else:
        # GET request: Pre-fill the forms with the current user’s data
        password_form = CustomPasswordChangeForm(request.user)

        # Debugging: Notify that we are rendering the page with pre-filled forms
        print("Rendering page with pre-filled forms.")

    # Render the profile update template with the forms
    return render(request, 'members/password_update.html', {
        'password_form': password_form
    })

@login_required
def profile(request):
    return render(request, 'members/profile.html')

# Excel Upload views- User List, Department, Designation

@login_required
@user_passes_test(is_super_user)
def export_user_list_excel(request):
    users = CustomUser.objects.select_related('approved_by').order_by('full_name')

    # Create workbook and worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = "User List"

    # Add header row
    headers = [
    'Sl. No.', 'Full Name', 'Email', 'Employee ID',
    'Designation', 'Department', 'Phone', 'Is Approved', 'Approved By', 'Approved At'
    ]
    ws.append(headers)

    # Add data rows
    for idx, user in enumerate(users, start=1):
        designation = str(user.designation) if user.designation else '—'
        if user.designation_other:
            designation += f" ({user.designation_other})"

        department = str(user.department) if user.department else '—'
        if user.department_other:
            department += f" ({user.department_other})"

        approved_at_str = '—'
        if user.approved_at:
            # Convert to local time and format nicely
            approved_at_str = localtime(user.approved_at).strftime('%d-%b-%Y %I:%M %p')

        ws.append([
            idx,
            user.full_name,
            user.email,
            user.employee_id or '—',
            designation,
            department,
            user.phone or '—',
            'Yes' if user.is_approved else 'No',
            user.approved_by.full_name if user.approved_by else '—',
            approved_at_str,
        ])

    # Generate dynamic filename with current timestamp
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f"user_list_{timestamp}.xlsx"

    # Prepare HTTP response with Excel content
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename={filename}'
    wb.save(response)

    return response

@login_required
@user_passes_test(is_super_user)
def export_department_list_excel(request):
    departments = Department.objects.select_related('created_by').order_by('name')

    # Create workbook and worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Department List"

    # Header row
    headers = [
        'Sl. No.',
        'Department Name',
        'Abbreviation',
        'Is Active',
        'Created At',
        'Created By'
    ]
    ws.append(headers)

    # Data rows
    for idx, dept in enumerate(departments, start=1):
        ws.append([
            idx,
            dept.name,
            dept.abbreviation,
            'Yes' if dept.is_active else 'No',
            dept.created_at.strftime('%Y-%m-%d %H:%M:%S') if dept.created_at else '',
            dept.created_by.full_name if dept.created_by else ''  # or dept.created_by.email
        ])

    # Filename with timestamp
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f"department_list_{timestamp}.xlsx"

    # Prepare Excel file for download
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename={filename}'
    wb.save(response)

    return response

@login_required
@user_passes_test(is_super_user)
def export_ward_list_excel(request):
    wards = Ward.objects.select_related('created_by').order_by('ward_name')

    wb = Workbook()
    ws = wb.active
    ws.title = "Ward List"

    headers = [
        'Sl. No.',
        'Ward Name',
        'Abbreviation',
        'Is Active',
        'Created At',
        'Created By'
    ]
    ws.append(headers)

    for idx, ward in enumerate(wards, start=1):
        ws.append([
            idx,
            ward.ward_name,
            ward.ward_abbre,
            'Yes' if ward.is_active else 'No',
            ward.created_at.strftime('%Y-%m-%d %H:%M:%S') if ward.created_at else '',
            ward.created_by.full_name if ward.created_by else ''
        ])

    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f"ward_list_{timestamp}.xlsx"

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename={filename}'
    wb.save(response)

    return response

@login_required
@user_passes_test(is_super_user)
def export_designation_list_excel(request):
    designations = Designation.objects.select_related('created_by').order_by('title')

    # Create workbook and worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Designation List"

    # Add header row
    headers = [
        'Sl. No.',
        'Designation Title',
        'Abbreviation',
        'Is Active',
        'Created At',
        'Created By'
    ]
    ws.append(headers)

    # Add data rows
    for idx, designation in enumerate(designations, start=1):
        ws.append([
            idx,
            designation.title,
            designation.abbreviation,
            'Yes' if designation.is_active else 'No',
            designation.created_at.strftime('%Y-%m-%d %H:%M:%S') if designation.created_at else '',
            designation.created_by.full_name if designation.created_by else ''
        ])

    # Generate dynamic filename with current timestamp
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f"designation_list_{timestamp}.xlsx"

    # Prepare HTTP response with Excel content
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename={filename}'
    wb.save(response)

    return response

@login_required
@user_passes_test(is_super_user)
def export_tag_list_excel(request):
    tags = CustomTag.objects.order_by('name')

    # Create workbook and worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Tag List"

    # Header row
    headers = [
        'Sl. No.',
        'Tag Name',
        'Abbreviation',
        'Value',
        'Tag Type',
        'Is Deleted'
    ]
    ws.append(headers)

    # Data rows
    for idx, tag in enumerate(tags, start=1):
        ws.append([
            idx,
            tag.name,
            tag.abbreviation,
            tag.value,
            tag.type,
            'Yes' if tag.is_deleted else 'No'
        ])

    # Filename with timestamp
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f"tag_list_{timestamp}.xlsx"

    # Prepare Excel file for download
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename={filename}'
    wb.save(response)

    return response


@login_required
@user_passes_test(is_super_user)
def download_user_template(request):
    wb = Workbook()
    ws = wb.active
    ws.title = "Users"

    # 1. Main header
    ws.append(['Email', 'Full Name', 'Is Approved', 'ID', 'Department', 'Designation', 'Phone'])

    # 2. Existing users
    for user in CustomUser.objects.select_related('department', 'designation').order_by('full_name'):
        ws.append([
            user.email,
            user.full_name,
            'Yes' if user.is_approved else 'No',
            user.employee_id,
            user.department.name if user.department else '',
            user.designation.title if user.designation else '',
            user.phone or ''
        ])

    # 3. Blank row
    ws.append([])

    # 4. Valid Departments
    ws.append(['Valid Departments'])
    for dept in Department.objects.filter(is_active=True).order_by('name'):
        ws.append([dept.name])

    # 5. Blank row
    ws.append([])

    # 6. Valid Designations
    ws.append(['Valid Designations'])
    for des in Designation.objects.filter(is_active=True).order_by('title'):
        ws.append([des.title])

    # 7. Return as Excel file
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    filename = f"user_template_{timezone.now().strftime('%Y-%m-%d_%H-%M-%S')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


@login_required
@user_passes_test(is_super_user)
def upload_user_template(request):
    form = UploadExcelForm(request.POST or None, request.FILES or None)

    if request.method == 'POST' and form.is_valid():
        file = form.cleaned_data['file']

        try:
            wb = openpyxl.load_workbook(file)
            ws = wb.active
        except Exception as e:
            messages.error(request, f"Error reading file: {e}")
            return redirect('members:upload_user_template')

        rows = list(ws.iter_rows(values_only=True))
        results = []
        updated_count = 0
        skipped_count = 0

        header = rows[0]
        for row in rows[1:]:
            if all((cell is None or str(cell).strip() == '') for cell in row):
                break

            label = str(row[0]).strip() if row[0] else ''
            if label == 'Valid Departments':
                break
            if not label:
                skipped_count += 1
                results.append({
                    'email': 'N/A',
                    'symbol': '❌',
                    'message': 'Missing Email',
                    'changes': [],
                    'success': False
                })
                continue

            try:
                user = CustomUser.objects.get(email__iexact=label)
            except CustomUser.DoesNotExist:
                skipped_count += 1
                results.append({
                    'email': label,
                    'symbol': '❌',
                    'message': 'User not found',
                    'changes': [],
                    'success': False
                })
                continue

            changes = []
            total_requests = 0
            successful = 0

            # Employee ID
            proposed = str(row[3]).strip() if row[3] else ''
            if proposed and proposed != (user.employee_id or ''):
                user.employee_id = proposed
                changes.append(f'ID: {user.employee_id} → {proposed}')
                successful += 1
                total_requests += 1

            # Department
            proposed = str(row[4]).strip() if row[4] else ''
            current = user.department.name if user.department else ''
            if proposed and proposed != current:
                try:
                    user.department = Department.objects.get(name__iexact=proposed)
                    changes.append(f'Department: {current} → {proposed}')
                    successful += 1
                except Department.DoesNotExist:
                    changes.append(f'Department: {current} → ❌ {proposed} (Invalid)')
                total_requests += 1

            # Designation
            proposed = str(row[5]).strip() if row[5] else ''
            current = user.designation.title if user.designation else ''
            if proposed and proposed != current:
                try:
                    user.designation = Designation.objects.get(title__iexact=proposed)
                    changes.append(f'Designation: {current} → {proposed}')
                    successful += 1
                except Designation.DoesNotExist:
                    changes.append(f'Designation: {current} → ❌ {proposed} (Invalid)')
                total_requests += 1

            # Phone
            proposed = str(row[6]).strip() if row[6] else ''
            if proposed and proposed != (user.phone or ''):
                user.phone = proposed
                changes.append(f'Phone: {user.phone} → {proposed}')
                successful += 1
                total_requests += 1

            if total_requests == 0:
                continue

            if successful > 0:
                user.save()
                updated_count += 1
                symbol = '✅' if successful == total_requests else '⚠️'
                message = 'Success' if successful == total_requests else 'Partial'
                flag = successful == total_requests
            else:
                symbol, message, flag = '❌', 'Failed', False
                skipped_count += 1

            results.append({
                'email': label,
                'symbol': symbol,
                'message': message,
                'changes': changes,
                'success': flag
            })

        if not results:
            messages.info(request, "Nothing to update.")
            return render(request, 'members/upload_template.html', {'form': form})

        return render(request, 'members/upload_template.html', {
            'form': form,
            'results': results,
            'summary': f'{updated_count} updated, {skipped_count} skipped.'
        })

    return render(request, 'members/upload_template.html', {'form': form})
